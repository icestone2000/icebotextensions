from __future__ import annotations

import os
import queue
import re
import shutil
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from config.logging import app_logger
from core.extension_manager import get_extensions_base_dir

# 打包在扩展目录里的台账模板（表头 + "例子："示例行）
TEMPLATE_FILENAME = "报修台账模板.xlsx"
# 工作表名包含该关键字时优先选用，否则用第一个工作表
SHEET_NAME_KEYWORD = "故障汇总"

# 台账列号（1 起）：序号/区域/路段/收费站/车道/故障状态/处理计时/故障描述/
# 故障报修时间/报修途径/报修人及电话/处理时间/故障原因/维护人员
COL_SEQ = 1
COL_STATION = 4
COL_LANE = 5
COL_DESCRIPTION = 8
COL_REPORT_TIME = 9
COL_CHANNEL = 10
COL_REPORTER = 11

REPORT_CHANNEL = "微信"

# worker 空转时等待新记录的超时（秒），同时也是 stop() 的响应粒度
QUEUE_WAIT_TIMEOUT_SEC = 1.0
# 信息不完整的报修等待补充的时长（秒），超时按现有信息写入台账
PENDING_TIMEOUT_SEC = 300.0
# 去重窗口（秒）：同一（台账+收费站+车道+报修人）在窗口内只记录一次
DEDUP_WINDOW_SEC = 1800.0
# 待补充记录监控线程的轮询间隔（秒）
PENDING_MONITOR_INTERVAL_SEC = 1.0

# 收费站视为"未提供"的取值
_UNKNOWN_STATIONS = ("", "未知", "unknown", "不详", "无")

_TIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d",
)


@dataclass
class RepairRecord:
    station: str
    lane: str
    description: str
    reporter: str
    report_time: datetime


@dataclass
class _PendingEntry:
    record: RepairRecord
    excel_path: str
    lane_required: bool
    deadline: float
    ask_count: int = 1


def parse_report_time(value: Optional[str]) -> datetime:
    """解析报修时间；缺省或解析失败时使用当前时间。"""
    s = str(value or "").strip()
    if s:
        for fmt in _TIME_FORMATS:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        app_logger.warning("[etc_repair_ledger] 无法解析报修时间 %r，使用当前时间", s)
    return datetime.now()


def resolve_excel_path(excel_file: str) -> str:
    """把工具入参解析为最终 .xlsx 绝对路径。

    - 相对名（不含目录）相对程序目录（get_extensions_base_dir）
    - .xls 后缀归一化为 .xlsx；无后缀补 .xlsx
    - 其他后缀视为非法
    """
    name = str(excel_file or "").strip().strip('"')
    if not name:
        raise ValueError("excel_file 不能为空")

    root, ext = os.path.splitext(name)
    ext_lower = ext.lower()
    if ext_lower in ("", ".xls"):
        name = root + ".xlsx"
    elif ext_lower != ".xlsx":
        raise ValueError(f"不支持的Excel文件后缀: {ext}")

    if not os.path.dirname(name):
        name = os.path.join(get_extensions_base_dir(), name)
    return os.path.normpath(os.path.abspath(name))


def station_missing(station: str) -> bool:
    return str(station or "").strip().lower() in _UNKNOWN_STATIONS


def _norm_station(station: str) -> str:
    s = "".join(str(station or "").split())
    for suf in ("收费站", "站"):
        if s.endswith(suf) and len(s) > len(suf):
            s = s[: -len(suf)]
            break
    return s


def _lanes_match(a: str, b: str) -> bool:
    """车道宽松匹配：都含数字比最后一个数字；否则按包含关系（LLM对同一报修
    的概括可能是 "入口"/"入口磅秤"、"出口84"/"84" 等变体）。"""
    a = "".join(str(a or "").split())
    b = "".join(str(b or "").split())
    da = re.findall(r"\d+", a)
    db = re.findall(r"\d+", b)
    if da and db:
        return da[-1].lstrip("0") == db[-1].lstrip("0")
    return a in b or b in a


def _get_template_path() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), TEMPLATE_FILENAME)


def _pick_worksheet(wb):
    for ws in wb.worksheets:
        if SHEET_NAME_KEYWORD in str(ws.title):
            return ws
    return wb.worksheets[0]


def _next_seq_and_row(ws) -> tuple:
    """返回 (下一个序号, 追加行号)。跳过表头与"例子："等非数字序号行。"""
    max_seq = 0
    last_used_row = 1
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, 15)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        last_used_row = row
        seq = values[COL_SEQ - 1]
        try:
            max_seq = max(max_seq, int(float(seq)))
        except (TypeError, ValueError):
            continue
    return max_seq + 1, last_used_row + 1


class RepairLedgerService:
    """报修台账写入服务。

    - 每个Excel文件一个队列 + 一个后台写入worker，enqueue() 只入队立即返回，
      worker 把积压记录批量写入（打开 -> 追加 -> 原子替换）。
    - submit() 在入队前做完整性检查：收费站必填，lane_required 时车道必填。
      信息不完整的记录进入待补充区，由监控线程在超时后按现有信息落账；
      期间同一报修人补齐信息则合并后落账（保留首次报修时间）。
    - 去重：同一（台账+收费站+车道+报修人）在 dedup_window_sec 内只记录一次。
    """

    def __init__(
        self,
        pending_timeout_sec: float = PENDING_TIMEOUT_SEC,
        dedup_window_sec: float = DEDUP_WINDOW_SEC,
    ):
        self._pending_timeout_sec = pending_timeout_sec
        self._dedup_window_sec = dedup_window_sec
        self._lock = threading.Lock()
        self._queues: Dict[str, queue.Queue] = {}
        self._threads: Dict[str, threading.Thread] = {}
        self._pending: Dict[Tuple[str, str, str], _PendingEntry] = {}
        # 近期已记录指纹：(excel_path, 站名归一, 车道原文归一, 报修人, 报修时间戳, 入库墙钟)
        self._recent: List[Tuple[str, str, str, str, float, float]] = []
        self._monitor_thread: Optional[threading.Thread] = None
        self._stop = threading.Event()

    # ------------------------------------------------------------------ 对外API
    def submit(
        self,
        excel_path: str,
        record: RepairRecord,
        lane_required: bool = False,
        group_key: str = "",
    ) -> Dict:
        """提交一条报修。返回 dict，status 取值：

        - awaiting_info: 信息不完整，已暂存等待补充（missing/already_asked/wait_seconds）
        - recorded: 已提交写入队列（merged=True 表示合并了此前的待补充记录）
        - duplicate: 去重窗口内的重复报修，已跳过
        """
        if self._stop.is_set():
            raise RuntimeError("服务已停止，无法提交报修记录")

        key = (excel_path, str(group_key or ""), record.reporter.strip())
        with self._lock:
            entry = self._pending.get(key)
            merged = False
            if entry is not None:
                record = self._merge_records(entry.record, record)
                lane_required = lane_required or entry.lane_required
                merged = True

            missing = self._missing_fields(record, lane_required)
            if missing:
                if entry is not None:
                    entry.record = record
                    entry.lane_required = lane_required
                    entry.ask_count += 1
                    already_asked = True
                else:
                    self._pending[key] = _PendingEntry(
                        record=record,
                        excel_path=excel_path,
                        lane_required=lane_required,
                        deadline=time.time() + self._pending_timeout_sec,
                    )
                    self._ensure_monitor_locked()
                    already_asked = False
                return {
                    "status": "awaiting_info",
                    "missing": missing,
                    "already_asked": already_asked,
                    "wait_seconds": int(self._pending_timeout_sec),
                }

            if entry is not None:
                self._pending.pop(key, None)

            pending_count = self._dedup_enqueue_locked(excel_path, record)
            if pending_count is None:
                return {"status": "duplicate"}

        return {"status": "recorded", "pending": pending_count, "merged": merged}

    def enqueue(self, excel_path: str, record: RepairRecord) -> Optional[int]:
        """入队写入（带去重，不做完整性检查）。重复报修返回 None。"""
        if self._stop.is_set():
            raise RuntimeError("服务已停止，无法提交报修记录")
        with self._lock:
            return self._dedup_enqueue_locked(excel_path, record)

    def stop(self, flush_timeout: float = 5.0) -> None:
        """停止服务；先把待补充记录与队列中的记录冲刷落账。"""
        with self._lock:
            pendings = list(self._pending.values())
            self._pending.clear()
            for entry in pendings:
                app_logger.info(
                    "[etc_repair_ledger] 服务停止，按现有信息写入待补充报修: %s/%s",
                    entry.record.station,
                    entry.record.reporter,
                )
                self._dedup_enqueue_locked(entry.excel_path, entry.record)
        self._stop.set()
        threads: List[threading.Thread] = []
        with self._lock:
            threads = list(self._threads.values())
            if self._monitor_thread is not None:
                threads.append(self._monitor_thread)
        for t in threads:
            t.join(timeout=flush_timeout)

    # ------------------------------------------------------------------ 完整性/去重
    @staticmethod
    def _missing_fields(record: RepairRecord, lane_required: bool) -> List[str]:
        missing = []
        if station_missing(record.station):
            missing.append("收费站")
        if lane_required and not str(record.lane or "").strip():
            missing.append("车道")
        return missing

    @staticmethod
    def _merge_records(old: RepairRecord, new: RepairRecord) -> RepairRecord:
        """合并待补充记录与新提交：新值非空则用新值，报修时间保留最早的。"""
        return RepairRecord(
            station=new.station if not station_missing(new.station) else old.station,
            lane=new.lane.strip() or old.lane,
            description=new.description.strip() or old.description,
            reporter=old.reporter,
            report_time=min(old.report_time, new.report_time),
        )

    def _dedup_enqueue_locked(self, excel_path: str, record: RepairRecord) -> Optional[int]:
        """去重后入队；重复返回 None。

        同一（台账+收费站+报修人）、车道宽松匹配、且报修时间相差不超过
        去重窗口的记录视为重复（按报修时间而非墙钟比较，兼容补处理历史消息）。
        """
        now = time.time()
        self._recent = [e for e in self._recent if now - e[5] <= self._dedup_window_sec]
        station = _norm_station(record.station)
        lane = "".join(str(record.lane or "").split())
        reporter = record.reporter.strip()
        report_ts = record.report_time.timestamp()
        for path, s, l, r, ts, _wall in self._recent:
            if (
                path == excel_path
                and s == station
                and r == reporter
                and _lanes_match(l, lane)
                and abs(report_ts - ts) <= self._dedup_window_sec
            ):
                return None
        self._recent.append((excel_path, station, lane, reporter, report_ts, now))
        return self._enqueue_locked(excel_path, record)

    # ------------------------------------------------------------------ 待补充监控
    def _ensure_monitor_locked(self) -> None:
        if self._monitor_thread is None or not self._monitor_thread.is_alive():
            self._monitor_thread = threading.Thread(
                target=self._monitor_pending,
                name="etc_repair_ledger:pending-monitor",
                daemon=True,
            )
            self._monitor_thread.start()

    def _monitor_pending(self) -> None:
        while not self._stop.wait(PENDING_MONITOR_INTERVAL_SEC):
            now = time.time()
            with self._lock:
                expired_keys = [k for k, e in self._pending.items() if e.deadline <= now]
                for key in expired_keys:
                    entry = self._pending.pop(key)
                    if self._dedup_enqueue_locked(entry.excel_path, entry.record) is None:
                        app_logger.info(
                            "[etc_repair_ledger] 待补充报修超时且与已记录重复，跳过: %s/%s",
                            entry.record.station,
                            entry.record.reporter,
                        )
                        continue
                    app_logger.info(
                        "[etc_repair_ledger] 待补充报修超时，按现有信息写入台账: 收费站=%s 车道=%s 报修人=%s",
                        entry.record.station or "未知",
                        entry.record.lane or "-",
                        entry.record.reporter,
                    )

    # ------------------------------------------------------------------ 队列/worker
    def _enqueue_locked(self, excel_path: str, record: RepairRecord) -> int:
        q = self._queues.get(excel_path)
        if q is None:
            q = queue.Queue()
            self._queues[excel_path] = q
        q.put(record)
        thread = self._threads.get(excel_path)
        if thread is None or not thread.is_alive():
            thread = threading.Thread(
                target=self._worker,
                args=(excel_path, q),
                name=f"etc_repair_ledger:{os.path.basename(excel_path)}",
                daemon=True,
            )
            self._threads[excel_path] = thread
            thread.start()
        return q.qsize()

    def _worker(self, excel_path: str, q: queue.Queue) -> None:
        while True:
            try:
                record = q.get(timeout=QUEUE_WAIT_TIMEOUT_SEC)
            except queue.Empty:
                if self._stop.is_set():
                    return
                continue
            batch: List[RepairRecord] = [record]
            while True:
                try:
                    batch.append(q.get_nowait())
                except queue.Empty:
                    break
            self._write_batch_with_retry(excel_path, batch)
            if self._stop.is_set() and q.empty():
                return

    def _write_batch_with_retry(self, excel_path: str, batch: List[RepairRecord]) -> None:
        for attempt in (1, 2):
            try:
                self._write_batch(excel_path, batch)
                app_logger.info(
                    "[etc_repair_ledger] 已写入 %d 条报修记录到 %s", len(batch), excel_path
                )
                return
            except Exception as exc:
                if attempt == 1:
                    app_logger.warning(
                        "[etc_repair_ledger] 写入 %s 失败，准备重试: %s", excel_path, exc
                    )
                else:
                    app_logger.exception(
                        "[etc_repair_ledger] 写入 %s 重试仍失败，丢弃 %d 条记录: %s",
                        excel_path,
                        len(batch),
                        exc,
                    )

    def _write_batch(self, excel_path: str, batch: List[RepairRecord]) -> None:
        self._ensure_file(excel_path)
        wb = load_workbook(excel_path)
        try:
            ws = _pick_worksheet(wb)
            seq, row = _next_seq_and_row(ws)
            for record in batch:
                ws.cell(row=row, column=COL_SEQ, value=seq)
                ws.cell(row=row, column=COL_STATION, value=record.station or "未知")
                if record.lane:
                    ws.cell(row=row, column=COL_LANE, value=record.lane)
                ws.cell(row=row, column=COL_DESCRIPTION, value=record.description)
                time_cell = ws.cell(row=row, column=COL_REPORT_TIME, value=record.report_time)
                time_cell.number_format = "yyyy/m/d hh:mm:ss"
                ws.cell(row=row, column=COL_CHANNEL, value=REPORT_CHANNEL)
                ws.cell(row=row, column=COL_REPORTER, value=record.reporter)
                seq += 1
                row += 1
            # 先写临时文件再原子替换，避免读取方看到写了一半的文件
            tmp_path = excel_path + ".tmp"
            wb.save(tmp_path)
        finally:
            wb.close()
        os.replace(tmp_path, excel_path)

    @staticmethod
    def _ensure_file(excel_path: str) -> None:
        if os.path.exists(excel_path):
            return
        parent = os.path.dirname(excel_path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        template = _get_template_path()
        if not os.path.exists(template):
            raise FileNotFoundError(f"台账模板不存在: {template}")
        shutil.copy(template, excel_path)
        app_logger.info("[etc_repair_ledger] 已从模板创建台账文件: %s", excel_path)
